from openpyxl import load_workbook
import fnmatch
import os
from tqdm import tqdm


class Member:
    def __init__(self, name='', dept='', dt='', email='', contents=[]):
        self.name = name
        self.dept = dept
        self.dt = dt
        self.email = email
        self.contents = contents

    def __str__(self):
        return f'{self.name}, {self.dept}, {self.dt}, {self.email}, {self.contents}'


def parse_line(line, key, defaultValue=''):
    if key in line:
        return line.split(':')[1].strip()
    else:
        return defaultValue


def parse_report(report_path: str) -> Member:
    m = Member()
    b_content = False
    with open(report_path, 'r', encoding='utf8') as f:
        for line in f:
            line = line.strip()
            if line.startswith('DEPT'):
                m.dept = parse_line(line, 'DEPT')
            elif line.startswith('NAME'):
                m.name = parse_line(line, 'NAME')
            elif line.startswith('EMAIL'):
                m.email = parse_line(line, 'EMAIL')
            elif line.startswith('DATE'):
                m.dt = parse_line(line, 'DATE')
            elif line.startswith('CONTENT'):
                b_content = True
                m.contents = []
            elif b_content and not line.startswith('==='):
                m.contents.append(line)

    return m


# Test
# m = parse_report('out/CX2_Steph Zhou2_s.zhou@example.com_WR_20200929_184000.txt')
# print(m)

def write_result(members: list, excel_file_path: str):
    wb = load_workbook(excel_file_path)
    sheet = wb.worksheets[0]
    for i, m in enumerate(members):
        col = i + 2
        sheet.cell(row=1, column=col).value = i + 1
        sheet.cell(row=2, column=col).value = m.name
        sheet.cell(row=3, column=col).value = m.dept
        sheet.cell(row=4, column=col).value = m.email
        sheet.cell(row=5, column=col).value = m.dt
        sheet.cell(row=6, column=col).value = '\n'.join(m.contents)

    wb.save(excel_file_path)


def collect_reports():
    report_dir = 'out'
    dest_file = 'WR


def collect_reports():
    report_dir = 'out'
    dest_file = 'WR_Summary.xlsx'
    reports = fnmatch.filter(os.listdir(report_dir), '*.txt')
    # members = []
    # for report in tqdm(reports):
    #     report_path = os.path.join(report_dir, report)
    #     m = parse_report(report_path)
    #     members.append(m)
    members = [parse_report(os.path.join(report_dir, report)) for report in tqdm(reports)]
    write_result(members, dest_file)
    print('Done!')


if __name__ == '__main__':
    collect_reports()
