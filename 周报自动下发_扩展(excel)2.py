import os
import fnmatch
from openpyxl import load_workbook
import openpyxl

def parse_line(line,key): #对每一行单独解析，返回需要的元素
    if key in line:
        return line.split(':')[1].strip()
def return_object(path):  # 对指定文件提取元素
    b_content = False
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()  # 去掉收尾空格或者换行符
            if line.startswith('DEPT'):
                dept=parse_line(line, 'DEPT')   # 定义函数parse_line来返回需要的值
            elif line.startswith('NAME'):
                name = parse_line(line, 'NAME')
            elif line.startswith('EMAIL'):
                email = parse_line(line, 'EMAIL')
            elif line.startswith('DATE'):
                dt = parse_line(line, 'DATE')
            elif line.startswith('CONTENT'):
                b_content = True  #读到CONTENT 时，标记为置位
                contents = []# 创建列表来接收contect的内容
            elif b_content and not line.startswith('==='): #标记为置位，且不是===开头
                contents.append(line)
        #HUB Bob Jiang bob@abc.com ['1.T0D01', '2.T0D02'] 20200930_184203
        return name,dept,email,dt,contents
def write2excel(path):
    col = 2
    for i in fnmatch.filter(os.listdir(path), '*.txt'):  #在指定目录中筛选出想要的文件，对每个文件遍历
        lst1=return_object(os.path.join(path,i))  #对每个文件遍历，使用return_object函数提取想要的元素
        print(lst1)
        wb=openpyxl.load_workbook('Report_Summary.xlsx')  #打开excel，存入
        sheet=wb.worksheets[0]
        sheet.cell(row=1, column=col).value=col-1    #row 代表行，conlumn 代表列
        sheet.cell(row=2, column=col).value = lst1[0]
        sheet.cell(row=3, column=col).value = lst1[1]
        sheet.cell(row=4, column=col).value = lst1[2]
        sheet.cell(row=5, column=col).value = lst1[3]
        sheet.cell(row=6, column=col).value = '\n'.join(lst1[4])
        col+=1
        wb.save('Report_Summary.xlsx')
write2excel('./report')
