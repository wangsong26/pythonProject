import os
import fnmatch
from openpyxl import load_workbook
import openpyxl
def return_object(path):  # 对指定文件提取元素
        lst=[]
        lst1=[]
        with open(path,'r',encoding='utf-8') as f:
            for line in f:
                line=line.strip()   #去掉收尾空格或者换行符
                lst.append(line)
             #lst=['DEPT: CX', 'NAME: Kevin Xin', 'EMAIL:xintq@qq.com', 'DATE: 20200930_184203', 'CONTENT:', '===========================', '1.T0D01', '2.T0D02']
            for i in lst:  #在列表中提取想要的元素
                if i == 'CONTENT:':
                    continue
                elif i.startswith('==='):
                    continue
                elif i.startswith('1.'):
                    lst1.append(i)
                elif i.startswith('2.'):
                    lst1.append(i)
                else:
                    lst1.append(i.split(':')[1])
        print(lst1)
        return lst1

def write2excel(path):
    col = 2
    for i in fnmatch.filter(os.listdir(path), '*.txt'):  #在指定目录中筛选出想要的文件，对每个文件遍历
        lst1=return_object(f'./report/{i}')  #对每个文件遍历，使用return_object函数提取想要的元素
        wb=openpyxl.load_workbook('Report_Summary.xlsx')  #打开excel，存入
        sheet=wb.worksheets[0]
        sheet.cell(row=1, column=col).value=col-1    #row 代表行，conlumn 代表列
        sheet.cell(row=2, column=col).value = lst1[1]
        sheet.cell(row=3, column=col).value = lst1[0]
        sheet.cell(row=4, column=col).value = lst1[2]
        sheet.cell(row=5, column=col).value = lst1[3]
        sheet.cell(row=6, column=col).value = '\n'.join(lst1[4:6])  # 使用换行符\n作为元素之间的分隔符
        col+=1
        wb.save('Report_Summary.xlsx')
write2excel('./report')






